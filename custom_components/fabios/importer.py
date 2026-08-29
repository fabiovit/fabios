from __future__ import annotations
import base64, csv, io, re
from datetime import date, datetime
from openpyxl import load_workbook
from pypdf import PdfReader
from .models import money

MONTHS_IT={"gennaio":1,"febbraio":2,"marzo":3,"aprile":4,"maggio":5,"giugno":6,"luglio":7,"agosto":8,"settembre":9,"ottobre":10,"novembre":11,"dicembre":12}
MONTH_PATTERN="|".join(MONTHS_IT)
ALIASES={
    "date":{"scadenza","data","date"},
    "description":{"azienda","descrizione","voce","esercente","merchant"},
    "total":{"importo totale","totale","importo","amount","total"},
    "shared":{"importo condiviso","condiviso","quota","shared","share"},
}

def _h(v): return re.sub(r"\s+"," ",str(v or "").strip().casefold())

def _num(v):
    if v is None or v=="": return None
    if isinstance(v,(int,float)): return float(v)
    t=str(v).replace("\xa0"," ").replace("€","").strip()
    t=re.sub(r"[^\d,.\-+]","",t)
    if not t:return None
    if "," in t and "." in t:
        t=t.replace(".","").replace(",",".") if t.rfind(",")>t.rfind(".") else t.replace(",","")
    else:t=t.replace(",",".")
    try:return float(t)
    except ValueError:return None

def _iso(v,year_hint=None):
    if isinstance(v,datetime):return v.date().isoformat()
    if isinstance(v,date):return v.isoformat()
    t=str(v or "").strip().casefold()
    for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%d/%m/%y"):
        try:return datetime.strptime(t,fmt).date().isoformat()
        except ValueError:pass
    for name,month in MONTHS_IT.items():
        if name in t:
            m=re.search(r"\b(20\d{2})\b",t)
            year=int(m.group(1)) if m else (year_hint or date.today().year)
            return f"{year:04d}-{month:02d}-01"
    return date.today().isoformat()

def _month_marker(description):
    text=re.sub(r"\s+"," ",str(description or "").strip())
    m=re.search(rf"\b({MONTH_PATTERN})(?:\s+(20\d{{2}}))?$",text,re.I)
    if not m:return None,text
    before=text[:m.start()].rstrip()
    if re.search(r"\bdi\s*$",before,re.I):return None,text
    marker=m.group(1).casefold()+((" "+m.group(2)) if m.group(2) else "")
    return marker,before

def _installment(description):
    m=re.search(r"\b(\d+)\s+di\s+(\d+)\b",str(description),re.I)
    if not m:return str(description).strip(),None,None
    cur,total=int(m.group(1)),int(m.group(2))
    clean=(str(description)[:m.start()]+str(description)[m.end():]).strip(" -–·")
    return clean or str(description).strip(),cur,total

def _cols(headers):
    n=[_h(x) for x in headers];out={}
    for key,aliases in ALIASES.items():
        for i,v in enumerate(n):
            if v in aliases:out[key]=i;break
    return out

def _table(rows,source):
    hi=None;cols=None
    for i,row in enumerate(rows[:50]):
        c=_cols(row)
        if {"description","total"}.issubset(c):hi=i;cols=c;break
    if hi is None:raise ValueError("Non trovo le colonne della tabella spese")

    parsed=[];current_month=None
    for row in rows[hi+1:]:
        def cell(k):
            idx=cols.get(k)
            return row[idx] if idx is not None and idx<len(row) else None

        desc=str(cell("description") or "").strip()
        total=_num(cell("total"));shared=_num(cell("shared"))
        if not desc or total is None:continue
        if desc.casefold() in {"totale","totali","subtotal","subtotale"}:continue
        if abs(total)<.005 and (shared is None or abs(shared)<.005):continue

        marker,clean=_month_marker(desc)
        if marker:
            current_month=marker
            desc=clean

        date_raw=str(cell("date") or "").strip() or (current_month or "")
        desc,ic,it=_installment(desc)
        parsed.append({
            "description":desc,"date_raw":date_raw,"total":money(total),
            "shared":money(shared) if shared is not None else None,
            "source":source,"installment_current":ic,"installment_total":it
        })

    if not parsed:raise ValueError("Nessuna spesa importabile trovata")
    return {"rows":parsed,"source":source}

def parse_csv(data):
    text=data.decode("utf-8-sig",errors="replace")
    try:dialect=csv.Sniffer().sniff(text[:4096],delimiters=";,\t")
    except csv.Error:
        dialect=csv.excel;dialect.delimiter=";"
    return _table(list(csv.reader(io.StringIO(text),dialect)),"csv")

def parse_xlsx(data):
    wb=load_workbook(io.BytesIO(data),read_only=True,data_only=True)
    for ws in wb.worksheets:
        rows=[list(r) for r in ws.iter_rows(values_only=True)]
        if any({"description","total"}.issubset(_cols(r)) for r in rows[:50]):
            out=_table(rows,"xlsx");out["sheet"]=ws.title;return out
    raise ValueError("Nessun foglio Excel contiene le colonne richieste")

def _pdf_rows(text):
    rows=[["Scadenza","Azienda","Importo Totale","Importo Condiviso"]]
    current_month=None
    for raw in text.splitlines():
        line=re.sub(r"\s+"," ",raw.replace("\xa0"," ")).strip()
        low=line.casefold()
        if not line or "spese casa condivise" in low or "importo totale" in low or low.startswith("totale "):continue

        # Richiede i centesimi per non confondere "8 di 12" con un importo.
        matches=list(re.finditer(r"[-+]?\d+[.,]\d{2}\s*€?",line))
        if not matches:continue

        if len(matches)>=2:
            tm,sm=matches[-2],matches[-1]
            total=_num(tm.group());shared=_num(sm.group());prefix=line[:tm.start()].strip()
        else:
            tm=matches[-1]
            total=_num(tm.group());shared=None;prefix=line[:tm.start()].strip()

        if total is None or abs(total)<.005 or not prefix:continue

        marker,clean=_month_marker(prefix)
        if marker:
            current_month=marker
            prefix=clean
        rows.append([current_month or "",prefix,total,shared])
    return rows

def parse_pdf(data):
    reader=PdfReader(io.BytesIO(data))
    text="\n".join((p.extract_text() or "") for p in reader.pages)
    if not text.strip():raise ValueError("Il PDF non contiene testo estraibile")
    out=_table(_pdf_rows(text),"pdf");out["pages"]=len(reader.pages);return out

def parse_upload(filename,content_b64):
    data=base64.b64decode(content_b64);name=filename.casefold()
    if name.endswith(".pdf"):return parse_pdf(data)
    if name.endswith(".xlsx"):return parse_xlsx(data)
    if name.endswith(".csv") or name.endswith(".txt"):return parse_csv(data)
    raise ValueError("Formato non supportato. Usa PDF, XLSX o CSV.")

def convert_preview_rows(rows,positive_payer,negative_payer,group_id,category,year_hint=None):
    out=[]
    for r in rows:
        signed=float(r["total"])
        payer=positive_payer if signed>=0 else negative_payer
        other=negative_payer if signed>=0 else positive_payer
        amount=money(abs(signed))
        shared=r.get("shared")

        # Importo Condiviso vuoto = spesa personale del pagatore.
        # Nessun debito viene creato verso l'altra persona.
        other_share=money(abs(float(shared))) if shared is not None else 0.0
        other_share=min(other_share,amount)
        payer_share=money(amount-other_share)

        out.append({
            "description":r["description"],"amount":amount,"paid_by":payer,
            "shares":{payer:payer_share,other:other_share},
            "date":_iso(r.get("date_raw"),year_hint),
            "group_id":group_id,"category":category or "Altro",
            "notes":f"Importato da {r.get('source','file')}",
            "installment_current":r.get("installment_current"),
            "installment_total":r.get("installment_total"),
        })
    return out
